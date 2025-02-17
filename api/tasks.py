import logging
from celery import shared_task
from openpyxl import load_workbook
from django.db import IntegrityError, transaction
from requests import RequestException
import requests
from .models import Conta, ContaContabil, ModelForecastCcusto, ModelForecastContabil, ModelForecastKeep
import pandas as pd
import numpy as np
from statsmodels.tsa.stattools import adfuller
from statsmodels.tsa.arima.model import ARIMA
from api.models import ModelForecastEmployment
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from forex_python.converter import CurrencyRates
import time
from .models import VariacaoDolar
from decimal import Decimal
import random
from datetime import date

@shared_task
def importar_xlsx_plano_conta(file_path):
    # Excluir todos os dados da tabela RawPlanoConta, mas manter a estrutura da tabela
    Conta.objects.all().delete()

    try:
        # Carregar o arquivo XLSX
        wb = load_workbook(file_path)
        sheet = wb.active  # Selecionando a primeira aba do arquivo

        # Iterando sobre as linhas da planilha, começando na segunda linha (presumindo que a primeira é o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                # Criar a conta no banco de dados
                Conta.objects.create(
                    cd_empresa=row[0],  # CD_EMPRESA
                    nr_nivel=row[1],    # NR_NIVEL
                    cd_conta_sup=row[2],  # CD_CONTA_SUP
                    ds_titulo_sup=row[3],  # DS_TITULO_SUP
                    cd_classif_contab=row[4],  # CD_CLASSIF_CONTAB
                    cd_conta_reduz=row[5],  # CD_CONTA_REDUZ
                    ds_conta=row[6],  # DS_CONTA
                )
            except IntegrityError:
                continue  # Ignora erros de integridade de duplicação
    except Exception as e:
        # Caso ocorra algum erro durante o processamento, logue ou envie a exceção para algum serviço de monitoramento
        print(f"Erro ao processar o arquivo {file_path}: {e}")


@shared_task
def importar_xlsx_conta_contabil(file_path):
    # Excluir todos os dados da tabela ContaContabil, mas manter a estrutura da tabela
    try:
        with transaction.atomic():  # Usar uma transação para garantir atomicidade
            ContaContabil.objects.all().delete()

            # Carregar o arquivo XLSX
            wb = load_workbook(file_path)
            sheet = wb.active  # Selecionando a primeira aba do arquivo

            # Criar uma lista de objetos para inserir em massa (bulk)
            contas_contabeis = []
            
            # Criar um DataFrame para carregar os dados
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(row)

            # Definindo as colunas
            columns = [
                'cd_empresa', 'cd_estabelecimento', 'classificacao_conta', 'cod_conta_contabil',
                'descricao_conta_contabil', 'cod_centro_custo', 'desc_centro_custo', 'ano', 
                'mes', 'valor_realizado'
            ]
            
            df = pd.DataFrame(data, columns=columns)
            
            # Agrupar os dados pelo cd_empresa, cd_estabelecimento, ano, mes e somar o valor_realizado
            df_grouped = df.groupby(
                ['cd_empresa', 'cd_estabelecimento', 'classificacao_conta', 'cod_conta_contabil', 
                 'descricao_conta_contabil', 'cod_centro_custo', 'desc_centro_custo', 'ano', 'mes'], 
                as_index=False
            )['valor_realizado'].sum()

            # Agora inserimos os dados agrupados no banco de dados
            for _, row in df_grouped.iterrows():
                try:
                    conta = ContaContabil(
                        cd_empresa=row['cd_empresa'],
                        cd_estabelecimento=row['cd_estabelecimento'],
                        classificacao_conta=row['classificacao_conta'],
                        cod_conta_contabil=row['cod_conta_contabil'],
                        descricao_conta_contabil=row['descricao_conta_contabil'],
                        cod_centro_custo=row['cod_centro_custo'],
                        desc_centro_custo=row['desc_centro_custo'],
                        ano=row['ano'],
                        mes=row['mes'],
                        valor_realizado=row['valor_realizado'],
                    )
                    contas_contabeis.append(conta)
                except IntegrityError:
                    continue  # Ignora erros de integridade de duplicação

            # Inserir as contas agrupadas em lote
            if contas_contabeis:
                ContaContabil.objects.bulk_create(contas_contabeis)

    except Exception as e:
        # Caso ocorra algum erro durante o processamento, logue ou envie a exceção para algum serviço de monitoramento
        print(f"Erro ao processar o arquivo {file_path}: {e}")



# Função para calcular a média mensal
def calcular_media_mensal(ticker_base, ticker_destino, ano, mes, max_retries=3):
    cr = CurrencyRates()
    datas = []
    taxas = []
    
    if mes == 12:
        mes_proximo = 1
        ano_proximo = ano + 1
    else:
        mes_proximo = mes + 1
        ano_proximo = ano

    # Calculando os dias do mês
    primeiro_dia = datetime(ano, mes, 1)
    ultimo_dia = datetime(ano_proximo, mes_proximo, 1) - timedelta(days=1)
    
    for dia in range((ultimo_dia - primeiro_dia).days + 1):
        data_atual = primeiro_dia + timedelta(days=dia)
        retries = 0
        while retries < max_retries:
            try:
                # Tentando obter a taxa de câmbio
                taxa = cr.get_rate(ticker_base, ticker_destino, data_atual)
                datas.append(data_atual)
                taxas.append(taxa)
                break  # Se a requisição foi bem-sucedida, sai do loop
            except RequestException as e:
                print(f"Erro ao buscar a taxa de câmbio para {data_atual}: {e}")
                retries += 1
                time.sleep(2)  # Aguardar 2 segundos antes de tentar novamente
                if retries == max_retries:
                    print(f"Falha ao buscar taxa de câmbio após {max_retries} tentativas para {data_atual}. Usando taxa de fallback.")
                    taxas.append(0.12)  # Usar taxa de fallback após falhas consecutivas
            except Exception as ex:
                print(f"Erro inesperado ao buscar a taxa de câmbio para {data_atual}: {ex}")
                taxas.append(0.12)  # Usar taxa de fallback
                break  # Interromper o loop se houver um erro inesperado
    
    if taxas:
        taxa_media = np.mean(taxas)  # Média das taxas de câmbio do mês
    else:
        print(f"Não foi possível calcular a média para o mês {mes}/{ano}. Usando taxa de fallback.")
        taxa_media = 0.12  # Valor de fallback para taxa média

    return taxa_media



# Função para calcular a variação percentual
def calcular_variacao(valor_antigo, valor_novo):
    return round(((valor_novo - valor_antigo) / valor_antigo) * 100, 2)

# Task para popular os dados de variação do dólar
@shared_task
# Função para calcular a taxa de variação média dos últimos 5 anos
def calcular_taxa_media_5_anos(ticker_base, ticker_destino, ano_atual, mes_atual):
    taxa_media_total = 0
    num_anos = 5  # Média dos últimos 5 anos
    
    for ano in range(ano_atual - num_anos + 1, ano_atual + 1):
        taxa_media_total += calcular_media_mensal(ticker_base, ticker_destino, ano, mes_atual)
    
    return taxa_media_total / num_anos

def obter_variacao_dolar(ano, mes, max_retries=3):
    cr = CurrencyRates()
    data_atual = f"{ano}-{mes:02d}-01"  # Primeiro dia do mês
    retries = 0
    while retries < max_retries:
        try:
            # Usando a API forex-python para obter a taxa de câmbio do dólar
            taxa_dolar = cr.get_rate('USD', 'BRL', data_atual)
            return taxa_dolar
        except Exception as e:
            print(f"Erro ao buscar taxa de câmbio para {data_atual}: {e}")
            retries += 1
            time.sleep(2)
    
    # Se falhar, retornar uma taxa de fallback
    print(f"Falha ao buscar a variação do dólar após {max_retries} tentativas. Usando taxa de fallback.")
    return 5.5  # Usando uma taxa de fallback de 5 BRL por USD

def obter_inflacao(ano, mes, max_retries=3):
    url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.433?formato=application/json"
    retries = 0
    while retries < max_retries:
        try:
            response = requests.get(url)
            response.raise_for_status()  # Verifica se houve erro na requisição
            inflacao_data = response.json()
            
            # Procura pela inflação do mês especificado
            for item in inflacao_data:
                data = item['data']
                if data.startswith(f'{ano}-{mes:02d}'):  # Formato 'AAAA-MM'
                    return float(item['valor']) / 100  # Convertendo para percentual
            # Caso não tenha encontrado, tenta novamente
            retries += 1
            time.sleep(2)  # Aguardar 2 segundos entre as tentativas
        except requests.exceptions.RequestException as e:
            print(f"Erro ao buscar dados da inflação: {e}")
            retries += 1
            time.sleep(2)
    
    # Se falhar, retornar uma taxa de fallback
    print(f"Falha ao buscar a inflação após {max_retries} tentativas. Usando taxa de fallback.")
    return 0.03  # Usando 3% de inflação como taxa de fallback

def ajustar_previsao(previsao, ano, mes):
    # Obter inflação e variação do dólar
    inflacao = obter_inflacao(ano, mes)
    variacao_dolar = obter_variacao_dolar(ano, mes)
    
    # Mostrar as taxas de inflação e variação do dólar
    print(f"Inflação para {mes}/{ano}: {inflacao * 100:.2f}%")
    print(f"Variação do Dólar para {mes}/{ano}: {variacao_dolar:.2f} BRL")
    
    # Ajustar a previsão com ambas as taxas
    previsao_ajustada = previsao * (1 + inflacao) * (1 + (variacao_dolar / 100))  # Ajusta pela inflação e pela variação do dólar
    return previsao_ajustada

@shared_task
def previsao_arima_task():
    # Função para testar estacionariedade usando o teste de Dickey-Fuller
    def testar_estacionariedade(serie):
        resultado = adfuller(serie)
        print(f"Estatística do Teste ADF: {resultado[0]}")
        print(f"P-valor: {resultado[1]}")
        if resultado[1] > 0.05:
            print("Série temporal não é estacionária. Realizando diferenciação...")
            return False
        else:
            print("Série temporal é estacionária.")
            return True

    # Função para corrigir valores nulos e anomalias
    def corrigir_dados(df):
        df['valor_realizado'] = df['valor_realizado'].interpolate(method='linear')
        mean = df['valor_realizado'].mean()
        std = df['valor_realizado'].std()
        df = df[df['valor_realizado'] < mean + 3*std]
        df = df[df['valor_realizado'] > mean - 3*std]
        return df

    # Função para treinar o modelo ARIMA e fazer previsões
    def previsao_arima():
        ModelForecastEmployment.objects.all().delete()
        data = ContaContabil.objects.all().values('cd_empresa', 'ano', 'mes', 'valor_realizado')
        df = pd.DataFrame(data)
        df['valor_realizado'] = df['valor_realizado'].astype(float)
        df = corrigir_dados(df)
        df['data'] = pd.to_datetime(df[['ano', 'mes']].astype(str).agg('-'.join, axis=1), format='%Y-%m')
        df.set_index('data', inplace=True)
        empresas = df['cd_empresa'].unique()
        
        for empresa in empresas:
            df_empresa = df[df['cd_empresa'] == empresa]

            if not testar_estacionariedade(df_empresa['valor_realizado']):
                df_empresa['valor_realizado'] = df_empresa['valor_realizado'].diff().dropna()
                if not testar_estacionariedade(df_empresa['valor_realizado']):
                    print(f"A série de {empresa} ainda não é estacionária após diferenciação.")
                    continue

            train = df_empresa['valor_realizado'][:-6]
            test = df_empresa['valor_realizado'][-6:]
            model = ARIMA(train, order=(5, 1, 0))
            model_fit = model.fit()
            forecast = model_fit.forecast(steps=6)

            print(f"Forecast for empresa {empresa}:")
            print(forecast)

            if isinstance(forecast, pd.Series):
                forecast_values = forecast.values
            else:
                forecast_values = forecast

            # Ajustando a previsão com a inflação
            for i in range(6):
                mes = df_empresa['mes'].max() + i
                ano = df_empresa['ano'].max()
                forecast_values[i] = ajustar_previsao(forecast_values[i], ano, mes)

            print(f"Previsões ajustadas para empresa {empresa}:")
            print(f"Mes 1: {forecast_values[0]}")
            print(f"Mes 2: {forecast_values[1]}")
            print(f"Mes 3: {forecast_values[2]}")
            print(f"Mes 4: {forecast_values[3]}")
            print(f"Mes 5: {forecast_values[4]}")
            print(f"Mes 6: {forecast_values[5]}")

            # Gravar as previsões na tabela
            model_forecast = ModelForecastEmployment.objects.create(
                cd_empresa=empresa,
                ano=df_empresa['ano'].max(),
                mes=df_empresa['mes'].max(),
                valor_previsto_mes_1=forecast_values[0],
                valor_previsto_mes_2=forecast_values[1],
                valor_previsto_mes_3=forecast_values[2],
                valor_previsto_mes_4=forecast_values[3],
                valor_previsto_mes_5=forecast_values[4],
                valor_previsto_mes_6=forecast_values[5],
            )

    previsao_arima()

@shared_task
def previsao_arima_keep_task():
    # Função para testar estacionariedade usando o teste de Dickey-Fuller
    def testar_estacionariedade(serie):
        resultado = adfuller(serie)
        print(f"Estatística do Teste ADF: {resultado[0]}")
        print(f"P-valor: {resultado[1]}")
        if resultado[1] > 0.05:
            print("Série temporal não é estacionária. Realizando diferenciação...")
            return False
        else:
            print("Série temporal é estacionária.")
            return True

    # Função para corrigir valores nulos e anomalias
    def corrigir_dados(df):
        df['valor_realizado'] = df['valor_realizado'].interpolate(method='linear')
        mean = df['valor_realizado'].mean()
        std = df['valor_realizado'].std()
        df = df[df['valor_realizado'] < mean + 3*std]
        df = df[df['valor_realizado'] > mean - 3*std]
        return df

    # Função para treinar o modelo ARIMA e fazer previsões

    def previsao_arima_keep():
        ModelForecastKeep.objects.all().delete()
        data = ContaContabil.objects.all().values('cd_empresa', 'cd_estabelecimento', 'ano', 'mes', 'valor_realizado')
        df = pd.DataFrame(data)
        df['valor_realizado'] = df['valor_realizado'].astype(float)
        df = corrigir_dados(df)
        df['data'] = pd.to_datetime(df[['ano', 'mes']].astype(str).agg('-'.join, axis=1), format='%Y-%m')
        df.set_index('data', inplace=True)
        
        # Agrupar por empresa e estabelecimento
        grupos = df.groupby(['cd_empresa', 'cd_estabelecimento'])

        # Set up logging
        logging.basicConfig(level=logging.INFO)
        logger = logging.getLogger(__name__)

        for (empresa, estabelecimento), df_empresa in grupos:
            try:
                if not testar_estacionariedade(df_empresa['valor_realizado']):
                    df_empresa['valor_realizado'] = df_empresa['valor_realizado'].diff().dropna()
                    if not testar_estacionariedade(df_empresa['valor_realizado']):
                        logger.warning(f"A série de {empresa} - {estabelecimento} ainda não é estacionária após diferenciação.")
                        continue

                train = df_empresa['valor_realizado'][:-6]
                test = df_empresa['valor_realizado'][-6:]
                model = ARIMA(train, order=(5, 1, 0))
                model_fit = model.fit()
                forecast = model_fit.forecast(steps=6)

                logger.info(f"Forecast for empresa {empresa} - estabelecimento {estabelecimento}:")
                logger.info(forecast)

                if isinstance(forecast, pd.Series):
                    forecast_values = forecast.values
                else:
                    forecast_values = forecast


                logger.info(f"Previsões ajustadas para empresa {empresa} - estabelecimento {estabelecimento}:")
                for i, val in enumerate(forecast_values):
                    logger.info(f"Mes {i+1}: {val}")

                # Gravar as previsões na tabela com cd_empresa e cd_estabelecimento
                model_forecast = ModelForecastKeep.objects.create(
                    cd_empresa=empresa,
                    cd_estabelecimento=estabelecimento,
                    ano=df_empresa['ano'].max(),
                    mes=df_empresa['mes'].max(),
                    valor_previsto_mes_1=forecast_values[0],
                    valor_previsto_mes_2=forecast_values[1],
                    valor_previsto_mes_3=forecast_values[2],
                    valor_previsto_mes_4=forecast_values[3],
                    valor_previsto_mes_5=forecast_values[4],
                    valor_previsto_mes_6=forecast_values[5],
                )
            except Exception as e:
                logger.error(f"Erro ao processar empresa {empresa} - estabelecimento {estabelecimento}: {str(e)}")

    previsao_arima_keep()


@shared_task
def previsao_arima_contabil_task():
    # Função para testar estacionariedade usando o teste de Dickey-Fuller
    def testar_estacionariedade(serie):
        resultado = adfuller(serie)
        print(f"Estatística do Teste ADF: {resultado[0]}")
        print(f"P-valor: {resultado[1]}")
        if resultado[1] > 0.05:
            print("Série temporal não é estacionária. Realizando diferenciação...")
            return False
        else:
            print("Série temporal é estacionária.")
            return True

    # Função para corrigir valores nulos e anomalias
    def corrigir_dados(df):
        df['valor_realizado'] = df['valor_realizado'].interpolate(method='linear')
        mean = df['valor_realizado'].mean()
        std = df['valor_realizado'].std()
        df = df[df['valor_realizado'] < mean + 3*std]
        df = df[df['valor_realizado'] > mean - 3*std]
        return df

    # Função para treinar o modelo ARIMA e fazer previsões

    def previsao_arima_contabil():
        ModelForecastContabil.objects.all().delete()
        data = ContaContabil.objects.all().values('cd_empresa', 'descricao_conta_contabil', 'ano', 'mes', 'valor_realizado')
        df = pd.DataFrame(data)
        df['valor_realizado'] = df['valor_realizado'].astype(float)

        # Garantir que 'valor_realizado' seja do tipo float, convertendo erros para NaN
        df['valor_realizado'] = pd.to_numeric(df['valor_realizado'], errors='coerce')

        # Identificar e corrigir 'Inf' (positivos e negativos) e 'NaN'
        df['valor_realizado'].replace([float('inf'), -float('inf')], float('nan'), inplace=True)

        # Substituir valores 'NaN' ou 'Inf' pela média da coluna específica para cada 'descricao_conta_contabil'
        df['valor_realizado'] = df.groupby('descricao_conta_contabil')['valor_realizado'].transform(lambda x: x.fillna(x.mean()))

        # Substituir valores zero pela média da coluna específica para cada 'descricao_conta_contabil'
        df['valor_realizado'] = df.groupby('descricao_conta_contabil')['valor_realizado'].transform(lambda x: x.replace(0, x.mean()))

        df = corrigir_dados(df)
        df['data'] = pd.to_datetime(df[['ano', 'mes']].astype(str).agg('-'.join, axis=1), format='%Y-%m')
        df.set_index('data', inplace=True)
        
        # Agrupar por empresa e estabelecimento
        grupos = df.groupby(['cd_empresa', 'descricao_conta_contabil'])

        # Set up logging
        logging.basicConfig(level=logging.INFO)
        logger = logging.getLogger(__name__)

        for (empresa, descricao_conta_contabil), df_empresa in grupos:
            try:
                if not testar_estacionariedade(df_empresa['valor_realizado']):
                    df_empresa['valor_realizado'] = df_empresa['valor_realizado'].diff().dropna()
                    if not testar_estacionariedade(df_empresa['valor_realizado']):
                        logger.warning(f"A série de {empresa} - {descricao_conta_contabil} ainda não é estacionária após diferenciação.")
                        continue

                train = df_empresa['valor_realizado'][:-6]
                test = df_empresa['valor_realizado'][-6:]
                model = ARIMA(train, order=(5, 1, 0))
                model_fit = model.fit()
                forecast = model_fit.forecast(steps=6)

                logger.info(f"Forecast for empresa {empresa} - conta Contabil {descricao_conta_contabil}:")
                logger.info(forecast)

                if isinstance(forecast, pd.Series):
                    forecast_values = forecast.values
                else:
                    forecast_values = forecast


                logger.info(f"Previsões ajustadas para empresa {empresa} - conta contabil {descricao_conta_contabil}:")
                for i, val in enumerate(forecast_values):
                    logger.info(f"Mes {i+1}: {val}")

                # Gravar as previsões na tabela com cd_empresa e descricao_conta_contabil
                model_forecast = ModelForecastContabil.objects.create(
                    cd_empresa=empresa,
                    descricao_conta_contabil=descricao_conta_contabil,
                    ano=df_empresa['ano'].max(),
                    mes=df_empresa['mes'].max(),
                    valor_previsto_mes_1=forecast_values[0],
                    valor_previsto_mes_2=forecast_values[1],
                    valor_previsto_mes_3=forecast_values[2],
                    valor_previsto_mes_4=forecast_values[3],
                    valor_previsto_mes_5=forecast_values[4],
                    valor_previsto_mes_6=forecast_values[5],
                )
            except Exception as e:
                logger.error(f"Erro ao processar empresa {empresa} - conta contabil {descricao_conta_contabil}: {str(e)}")

    previsao_arima_contabil()



@shared_task
def previsao_arima_ccusto_task():
    # Função para testar estacionariedade usando o teste de Dickey-Fuller
    def testar_estacionariedade(serie):
        resultado = adfuller(serie)
        print(f"Estatística do Teste ADF: {resultado[0]}")
        print(f"P-valor: {resultado[1]}")
        if resultado[1] > 0.05:
            print("Série temporal não é estacionária. Realizando diferenciação...")
            return False
        else:
            print("Série temporal é estacionária.")
            return True

    # Função para corrigir valores nulos e anomalias
    def corrigir_dados(df):
        df['valor_realizado'] = df['valor_realizado'].interpolate(method='linear')
        mean = df['valor_realizado'].mean()
        std = df['valor_realizado'].std()
        df = df[df['valor_realizado'] < mean + 3*std]
        df = df[df['valor_realizado'] > mean - 3*std]
        return df



    def calcular_valores_planejados_por_mes(df):
        """
        Calcula os valores planejados baseados na média do valor realizado por mês.
        
        :param df: DataFrame contendo os dados históricos de valor realizado.
        :return: Dicionário com os valores planejados para os próximos 6 meses.
        """
        valores_planejados = {}
        mes_atual = df['mes'].max()  # Último mês nos dados

        for i in range(1, 7):  # Para os próximos 6 meses
            mes_previsto = (mes_atual + i - 1) % 12 + 1  # Calcula o mês previsto (1 a 12)
            media_mes = df[df['mes'] == mes_previsto]['valor_realizado'].mean()
            valores_planejados[f'valor_planejado_mes_{i}'] = media_mes

        return valores_planejados

    def previsao_arima_ccusto():
        ModelForecastCcusto.objects.all().delete()
        data = ContaContabil.objects.all().values('cd_empresa', 'descricao_conta_contabil','desc_centro_custo', 'ano', 'mes', 'valor_realizado')
        df = pd.DataFrame(data)
        df['valor_realizado'] = df['valor_realizado'].astype(float)

        # Garantir que 'valor_realizado' seja do tipo float, convertendo erros para NaN
        df['valor_realizado'] = pd.to_numeric(df['valor_realizado'], errors='coerce')

        # Identificar e corrigir 'Inf' (positivos e negativos) e 'NaN'
        df['valor_realizado'].replace([float('inf'), -float('inf')], float('nan'), inplace=True)

        # Substituir valores 'NaN' ou 'Inf' pela média da coluna específica para cada 'descricao_conta_contabil'
        df['valor_realizado'] = df.groupby('descricao_conta_contabil')['valor_realizado'].transform(lambda x: x.fillna(x.mean()))

        # Substituir valores zero pela média da coluna específica para cada 'descricao_conta_contabil'
        df['valor_realizado'] = df.groupby('descricao_conta_contabil')['valor_realizado'].transform(lambda x: x.replace(0, x.mean()))


        df = corrigir_dados(df)
        df['data'] = pd.to_datetime(df[['ano', 'mes']].astype(str).agg('-'.join, axis=1), format='%Y-%m')
        df.set_index('data', inplace=True)
        
        # Agrupar por empresa e estabelecimento
        grupos = df.groupby(['cd_empresa', 'descricao_conta_contabil', 'desc_centro_custo'])

        # Set up logging
        logging.basicConfig(level=logging.INFO)
        logger = logging.getLogger(__name__)

        for (empresa, descricao_conta_contabil, desc_centro_custo), df_empresa in grupos:
            try:
                if not testar_estacionariedade(df_empresa['valor_realizado']):
                    df_empresa['valor_realizado'] = df_empresa['valor_realizado'].diff().dropna()
                    if not testar_estacionariedade(df_empresa['valor_realizado']):
                        logger.warning(f"A série de {empresa} - {descricao_conta_contabil} - {desc_centro_custo} ainda não é estacionária após diferenciação.")
                        continue

                train = df_empresa['valor_realizado'][:-6]
                test = df_empresa['valor_realizado'][-6:]
                model = ARIMA(train, order=(5, 1, 0))
                model_fit = model.fit()
                forecast = model_fit.forecast(steps=6)

                logger.info(f"Forecast for empresa {empresa} - conta Contabil {descricao_conta_contabil} - CC {desc_centro_custo}:")
                logger.info(forecast)

                if isinstance(forecast, pd.Series):
                    forecast_values = forecast.values
                else:
                    forecast_values = forecast

                logger.info(f"Previsões ajustadas para empresa {empresa} - conta contabil {descricao_conta_contabil} - CC {desc_centro_custo}:")
                for i, val in enumerate(forecast_values):
                    logger.info(f"Mês {i+1}: {val}")

                # Calcular os valores planejados com base na média por mês
                valores_planejados = calcular_valores_planejados_por_mes(df_empresa)

                # Gravar as previsões na tabela com cd_empresa e descricao_conta_contabil
                ModelForecastCcusto.objects.create(
                    cd_empresa=empresa,
                    descricao_conta_contabil=descricao_conta_contabil,
                    desc_centro_custo=desc_centro_custo,
                    ano=df_empresa['ano'].max(),
                    mes=df_empresa['mes'].max(),
                    valor_previsto_mes_1=forecast_values[0],
                    valor_previsto_mes_2=forecast_values[1],
                    valor_previsto_mes_3=forecast_values[2],
                    valor_previsto_mes_4=forecast_values[3],
                    valor_previsto_mes_5=forecast_values[4],
                    valor_previsto_mes_6=forecast_values[5],
                    valor_planejado_mes_1=valores_planejados['valor_planejado_mes_1'],
                    valor_planejado_mes_2=valores_planejados['valor_planejado_mes_2'],
                    valor_planejado_mes_3=valores_planejados['valor_planejado_mes_3'],
                    valor_planejado_mes_4=valores_planejados['valor_planejado_mes_4'],
                    valor_planejado_mes_5=valores_planejados['valor_planejado_mes_5'],
                    valor_planejado_mes_6=valores_planejados['valor_planejado_mes_6'],
                )
            except Exception as e:
                logger.error(f"Erro ao processar empresa {empresa} - conta contabil {descricao_conta_contabil} - CC {desc_centro_custo}: {str(e)}")


    previsao_arima_ccusto()